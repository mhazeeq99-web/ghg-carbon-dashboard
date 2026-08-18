import openpyxl

SCOPE1 = "migration/GHG DATA_SCOPE 1.xlsx"
SCOPE2 = "migration/GHG DATA_SCOPE 2.xlsx"

MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

records = []
warnings = []


def read_month_table(
    ws,
    slug,
    start_row,
    locations,
):
    """
    Read exactly the 12-month activity table.

    start_row = January row.
    """

    for offset in range(12):
        row = start_row + offset

        month_name = ws.cell(row, 2).value

        if month_name not in MONTHS:
            raise ValueError(
                f"Expected month at row {row}, "
                f"found {month_name!r} in {ws.title}"
            )

        month = MONTHS[month_name]

        for location, year_columns in locations.items():

            for year, column in year_columns.items():

                value = ws.cell(row, column).value

                # Blank = no record
                if value is None:
                    continue

                # Excel errors
                if isinstance(value, str):

                    if value.startswith("#"):
                        warnings.append(
                            f"{slug} | {location} | {year} | "
                            f"{month_name} | {value}"
                        )
                        continue

                    raise ValueError(
                        f"Unexpected text value: "
                        f"{slug} | {location} | {year} | "
                        f"{month_name} | {value}"
                    )

                if not isinstance(value, (int, float)):
                    raise ValueError(
                        f"Unexpected value: "
                        f"{slug} | {location} | {year} | "
                        f"{month_name} | {value}"
                    )

                records.append({
                    "slug": slug,
                    "location": location,
                    "year": year,
                    "month": month,
                    "quantity": float(value),
                })


# ============================================================
# SCOPE 1
# ============================================================

wb = openpyxl.load_workbook(
    SCOPE1,
    data_only=True
)


read_month_table(
    wb["LPG 14kg"],
    "lpg-14kg",
    7,
    {
        "Tago": {
            2022: 9,
            2023: 10,
            2024: 11,
            2025: 12,
            2026: 13,
        },
        "KIP": {
            2022: 14,
            2023: 15,
            2024: 16,
            2025: 17,
            2026: 18,
        },
    },
)


read_month_table(
    wb["LPG 50kg"],
    "lpg-50kg",
    7,
    {
        "Tago": {
            2022: 3,
            2023: 4,
            2024: 5,
            2025: 6,
            2026: 7,
        },
    },
)


read_month_table(
    wb["Diesel"],
    "diesel",
    7,
    {
        "Tago": {
            2022: 3,
            2023: 4,
            2024: 5,
            2025: 6,
            2026: 7,
        },
    },
)


read_month_table(
    wb["Petrol"],
    "petrol",
    7,
    {
        "Tago": {
            2022: 3,
            2023: 4,
            2024: 5,
            2025: 6,
            2026: 7,
        },
    },
)


# ============================================================
# SCOPE 2
# ============================================================

wb = openpyxl.load_workbook(
    SCOPE2,
    data_only=True
)


read_month_table(
    wb["Electricity"],
    "electricity",
    7,
    {
        "Tago": {
            2022: 9,
            2023: 10,
            2024: 11,
            2025: 12,
            2026: 13,
        },
        "KIP": {
            2022: 14,
            2023: 15,
            2024: 16,
            2025: 17,
            2026: 18,
        },
    },
)


# ============================================================
# PREVIEW
# ============================================================

print()
print("=" * 70)
print("GHG MIGRATION PREVIEW")
print("=" * 70)

print(f"Total records: {len(records)}")

print()
print("Records by parameter / location:")

counts = {}

for r in records:
    key = (r["slug"], r["location"])
    counts[key] = counts.get(key, 0) + 1

for (slug, location), count in sorted(counts.items()):
    print(
        f"  {slug:12} | "
        f"{location:5} | "
        f"{count:3} records"
    )


print()
print("Warnings:")

if warnings:
    for warning in warnings:
        print("  WARNING:", warning)
else:
    print("  None")


print()
print("Sample records:")

for r in records[:20]:
    print(
        f"  {r['slug']:12} | "
        f"{r['location']:5} | "
        f"{r['year']} | "
        f"{r['month']:02d} | "
        f"{r['quantity']}"
    )


print()
print("=" * 70)
print("DRY RUN ONLY")
print("NOTHING WAS INSERTED INTO NEON")
print("=" * 70)
