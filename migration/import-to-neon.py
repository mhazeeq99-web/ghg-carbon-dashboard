import os
import openpyxl
import psycopg2
from dotenv import load_dotenv

load_dotenv(".env.local")

DATABASE_URL = os.environ["DATABASE_URL"]

SCOPE1 = "migration/GHG DATA_SCOPE 1.xlsx"
SCOPE2 = "migration/GHG DATA_SCOPE 2.xlsx"

MONTHS = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

records = []


def read_month_table(ws, slug, start_row, locations):
    for offset in range(12):
        row = start_row + offset
        month_name = ws.cell(row, 2).value

        if month_name not in MONTHS:
            raise ValueError(
                f"Expected month at row {row}, "
                f"found {month_name!r} in {ws.title}"
            )

        month = MONTHS[month_name]

        for location, years in locations.items():
            for year, column in years.items():
                value = ws.cell(row, column).value

                if value is None:
                    continue

                if isinstance(value, str):
                    if value.startswith("#"):
                        continue

                    raise ValueError(
                        f"Unexpected text: "
                        f"{slug} / {location} / {year} / "
                        f"{month_name} / {value}"
                    )

                if not isinstance(value, (int, float)):
                    raise ValueError(
                        f"Unexpected value: "
                        f"{slug} / {location} / {year} / "
                        f"{month_name} / {value}"
                    )

                records.append({
                    "slug": slug,
                    "location": location,
                    "year": year,
                    "month": month,
                    "quantity": float(value),
                })


# ============================================================
# READ SCOPE 1
# ============================================================

wb = openpyxl.load_workbook(SCOPE1, data_only=True)

read_month_table(
    wb["LPG 14kg"],
    "lpg-14kg",
    7,
    {
        "Tago": {2022: 9, 2023: 10, 2024: 11, 2025: 12, 2026: 13},
        "KIP":  {2022: 14, 2023: 15, 2024: 16, 2025: 17, 2026: 18},
    },
)

read_month_table(
    wb["LPG 50kg"],
    "lpg-50kg",
    7,
    {
        "Tago": {2022: 3, 2023: 4, 2024: 5, 2025: 6, 2026: 7},
    },
)

read_month_table(
    wb["Diesel"],
    "diesel",
    7,
    {
        "Tago": {2022: 3, 2023: 4, 2024: 5, 2025: 6, 2026: 7},
    },
)

read_month_table(
    wb["Petrol"],
    "petrol",
    7,
    {
        "Tago": {2022: 3, 2023: 4, 2024: 5, 2025: 6, 2026: 7},
    },
)


# ============================================================
# READ SCOPE 2
# ============================================================

wb = openpyxl.load_workbook(SCOPE2, data_only=True)

read_month_table(
    wb["Electricity"],
    "electricity",
    7,
    {
        "Tago": {2022: 9, 2023: 10, 2024: 11, 2025: 12, 2026: 13},
        "KIP":  {2022: 14, 2023: 15, 2024: 16, 2025: 17, 2026: 18},
    },
)


# ============================================================
# IMPORT TO NEON
# ============================================================

print()
print("=" * 70)
print("NEON GHG DATA MIGRATION")
print("=" * 70)
print(f"Records prepared: {len(records)}")

conn = psycopg2.connect(DATABASE_URL)

try:
    cur = conn.cursor()

    # Make sure all expected parameters exist
    expected_parameters = {
        "lpg-14kg",
        "lpg-50kg",
        "diesel",
        "petrol",
        "electricity",
    }

    cur.execute(
        "SELECT slug FROM parameters"
    )

    existing_parameters = {
        row[0] for row in cur.fetchall()
    }

    missing = expected_parameters - existing_parameters

    if missing:
        raise RuntimeError(
            f"Missing parameters in Neon: {sorted(missing)}"
        )

    # Make sure all locations exist
    cur.execute(
        "SELECT name FROM locations"
    )

    existing_locations = {
        row[0] for row in cur.fetchall()
    }

    required_locations = {
        r["location"] for r in records
    }

    missing_locations = required_locations - existing_locations

    if missing_locations:
        raise RuntimeError(
            f"Missing locations in Neon: {sorted(missing_locations)}"
        )

    inserted = 0

    for r in records:

        cur.execute(
            """
            INSERT INTO activity_data (
                parameter_id,
                location_id,
                year,
                month,
                quantity
            )
            SELECT
                p.id,
                l.id,
                %s,
                %s,
                %s
            FROM parameters p
            CROSS JOIN locations l
            WHERE p.slug = %s
              AND l.name = %s

            ON CONFLICT (
                parameter_id,
                location_id,
                year,
                month
            )
            DO UPDATE SET
                quantity = EXCLUDED.quantity,
                updated_at = NOW()
            """,
            (
                r["year"],
                r["month"],
                r["quantity"],
                r["slug"],
                r["location"],
            ),
        )

        inserted += 1

    conn.commit()

    print()
    print(f"Successfully migrated: {inserted} records")
    print()
    print("NEON MIGRATION COMPLETE")
    print("=" * 70)

except Exception as e:

    conn.rollback()

    print()
    print("MIGRATION FAILED")
    print(str(e))
    print("All changes were rolled back.")

    raise

finally:
    conn.close()